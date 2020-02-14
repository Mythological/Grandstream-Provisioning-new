# -*- coding: utf-8 -*-
# !/usr/bin/env python

from openpyxl import load_workbook
import pymysql
from credentials import *

conn = pymysql.connect(host,user,password, db)

x = conn.cursor()

wb = load_workbook(excel, data_only=True)
sheet = wb['Sheet']

PhonesDict = {}

for row in sheet.iter_rows():
     name = row[2].value
     extension = row[1].value
     secret = row[8].value
     callgroup = row[4].value
     pickup = row[5].value
     mac = row[3].value
     ring = row[6].value
     email = row[7].value
     ip3 = row[9].value
     ip4 = row[10].value
     model = row[11].value
     DID = row[12].value
     print(DID)
     try:
          x.execute("INSERT INTO users(name,ext,secret,callGroup,pickup,mac,ring,email,ip3,ip4,model,DID) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
          (name, extension, secret, callgroup, pickup, mac, ring, email, ip3, ip4, model, DID))
     except pymysql.err.IntegrityError:
          x.execute(
               'UPDATE users SET name=%s, secret=%s, callGroup=%s, pickup=%s, mac=%s, ring=%s, email=%s, ip3=%s, ip4=%s, model=%s, DID=%s WHERE ext=%s ',
               (name, secret, callgroup, pickup, mac, ring, email, ip3, ip4, model, DID, extension))
     conn.commit()
conn.close()
